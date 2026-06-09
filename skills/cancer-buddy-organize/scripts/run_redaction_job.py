#!/usr/bin/env python3
"""
Deterministic pre-persist redaction job for cancer-buddy-organize.

This script consumes redaction_manifest_v2. It never performs OCR. The LLM has
already supplied redacted Markdown plus PII region metadata; this script only
applies those coordinates/structure locators, writes redacted candidates, and
commits them after a separate LLM QA report passes.

Usage:
    python3 run_redaction_job.py prepare <patient_dir>
    python3 run_redaction_job.py prepare --manifest <path/to/redaction_manifest.json>
    python3 run_redaction_job.py commit <patient_dir> --qa-report <llm_redaction_qa.json>

Exit codes:
    0  success
    1  at least one file failed/blocked or commit was rejected by QA
    2  invocation or manifest/status error
"""
from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

MANIFEST_NAME = "redaction_manifest.json"
STATUS_NAME = "redaction_status.json"
SOURCE_STATUS_NAME = "source_redaction_status.json"
SOURCE_INVENTORY_NAME = "source_inventory.json"
MASK = "[PII_MASKED]"

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".webp", ".bmp", ".heic", ".heif"}
TEXT_EXTS = {".txt", ".csv", ".md", ".html", ".htm"}
STRATEGY_BY_KIND = {
    "image": "llm_region_image",
    "pdf": "llm_region_pdf",
    "docx": "llm_structured_docx",
    "xlsx": "llm_structured_sheet",
    "csv": "llm_text_rewrite",
    "txt": "llm_text_rewrite",
    "html": "llm_text_rewrite",
    "md": "llm_text_rewrite",
    "archive": "archive_rebuild",
    "other": "blocked_unsupported",
}


def now_iso() -> str:
    return _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")


def read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, doc: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_rel(path: str) -> str:
    if not path or path.startswith("/") or ".." in Path(path).parts:
        raise ValueError(f"unsafe patient-dir-relative path: {path!r}")
    return path


def resolve_manifest(args) -> tuple[Path, dict, Path]:
    if args.manifest:
        manifest_path = Path(args.manifest).resolve()
    else:
        patient_dir_arg = getattr(args, "patient_dir", None)
        if not patient_dir_arg:
            raise ValueError("patient_dir or --manifest is required")
        manifest_path = Path(patient_dir_arg).resolve() / MANIFEST_NAME
    manifest = read_json(manifest_path)
    if manifest.get("schema") != "redaction_manifest_v2":
        raise ValueError(f"{manifest_path}: expected schema redaction_manifest_v2")
    patient_dir = Path(manifest.get("patient_dir") or manifest_path.parent).resolve()
    return manifest_path, manifest, patient_dir


def load_status(patient_dir: Path) -> dict[str, dict]:
    path = patient_dir / STATUS_NAME
    if not path.is_file():
        return {}
    try:
        doc = read_json(path)
        if doc.get("schema") != "redaction_status_v2":
            return {}
        return {e["id"]: e for e in doc.get("files", []) if isinstance(e, dict) and e.get("id")}
    except Exception:
        return {}


def write_status(patient_dir: Path, entries: list[dict]) -> None:
    summary = {"total": len(entries), "pending": 0, "redacted_pending_qa": 0, "done": 0, "failed": 0, "blocked": 0}
    for e in entries:
        st = e.get("status", "pending")
        if st in summary:
            summary[st] += 1
    write_json(
        patient_dir / STATUS_NAME,
        {
            "schema": "redaction_status_v2",
            "patient_dir": str(patient_dir),
            "updated_at": now_iso(),
            "summary": summary,
            "files": entries,
        },
    )


def empty_entry(entry: dict) -> dict:
    return {
        "id": entry["id"],
        "source_id": entry["source_id"],
        "status": "pending",
        "redacted_candidate_path": None,
        "redacted_path": None,
        "coverage_passed": None,
        "llm_qa_passed": None,
        "qa_report_id": None,
        "qa_passed": None,
        "original_deleted": None,
        "reason": None,
    }


def candidate_rel(entry: dict, source_abs: Path) -> str:
    existing = entry.get("redacted_candidate_path")
    if existing:
        return safe_rel(existing)
    suffix = source_abs.suffix
    if source_abs.suffix.lower() in {".heic", ".heif"}:
        suffix = ".jpg"
    return f".redaction_candidates/{entry['id']}_{source_abs.stem}_redacted{suffix}"


def bbox_to_px(locator: dict, width: int, height: int, normalized: bool) -> tuple[int, int, int, int]:
    box_w = locator.get("w", locator.get("width"))
    box_h = locator.get("h", locator.get("height"))
    if box_w is None or box_h is None:
        raise ValueError(f"bbox locator requires w/width and h/height: {locator!r}")
    if normalized:
        x = float(locator["x"]) * width
        y = float(locator["y"]) * height
        w = float(box_w) * width
        h = float(box_h) * height
    else:
        x = float(locator["x"])
        y = float(locator["y"])
        w = float(box_w)
        h = float(box_h)
    x1 = max(0, min(width, int(round(x))))
    y1 = max(0, min(height, int(round(y))))
    x2 = max(0, min(width, int(round(x + w))))
    y2 = max(0, min(height, int(round(y + h))))
    if x2 <= x1 or y2 <= y1:
        raise ValueError(f"empty bbox after clipping: {locator!r}")
    return x1, y1, x2, y2


def quad_to_px(points: list, width: int, height: int, normalized: bool) -> list[tuple[int, int]]:
    if len(points) < 3:
        raise ValueError("quad locator needs at least 3 points")
    out = []
    for p in points:
        x = float(p[0]) * width if normalized else float(p[0])
        y = float(p[1]) * height if normalized else float(p[1])
        out.append((max(0, min(width, int(round(x)))), max(0, min(height, int(round(y))))))
    return out


def open_image(path: Path):
    from PIL import Image, ImageOps

    try:
        img = Image.open(path)
        return ImageOps.exif_transpose(img).convert("RGB")
    except Exception:
        if path.suffix.lower() not in {".heic", ".heif"} or not shutil.which("sips"):
            raise
        with tempfile.TemporaryDirectory(prefix="cb_heic_") as td:
            tmp = Path(td) / f"{path.stem}.jpg"
            subprocess.run(["sips", "-s", "format", "jpeg", str(path), "--out", str(tmp)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            img = Image.open(tmp)
            return ImageOps.exif_transpose(img).convert("RGB")


def redact_image(source: Path, out: Path, regions: list[dict], page: int | None = None) -> int:
    from PIL import ImageDraw

    img = open_image(source)
    width, height = img.size
    draw = ImageDraw.Draw(img)
    applied = 0
    for region in regions:
        locator_type = region["locator_type"]
        locator = region["locator"]
        region_page = locator.get("page")
        if page is not None and region_page is not None and int(region_page) != page:
            continue
        if locator_type in {"normalized_bbox", "page_normalized_bbox"}:
            draw.rectangle(bbox_to_px(locator, width, height, True), fill="black")
            applied += 1
        elif locator_type == "pixel_bbox":
            draw.rectangle(bbox_to_px(locator, width, height, False), fill="black")
            applied += 1
        elif locator_type in {"normalized_quad", "page_normalized_quad"}:
            draw.polygon(quad_to_px(locator["points"], width, height, True), fill="black")
            applied += 1
    if applied == 0:
        raise ValueError("no image/PDF coordinate regions applied")
    out.parent.mkdir(parents=True, exist_ok=True)
    save_kwargs = {}
    if out.suffix.lower() in {".jpg", ".jpeg"}:
        save_kwargs["quality"] = 95
    img.save(out, **save_kwargs)
    return applied


def redact_pdf(source: Path, out: Path, regions: list[dict], adapter_frame: dict) -> int:
    if not shutil.which("pdftoppm"):
        raise RuntimeError("pdftoppm is required for PDF region redaction")
    from PIL import Image
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas

    dpi = int(adapter_frame.get("dpi") or 150)
    with tempfile.TemporaryDirectory(prefix="cb_pdf_") as td:
        prefix = Path(td) / "page"
        subprocess.run(["pdftoppm", "-r", str(dpi), "-png", str(source), str(prefix)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        pages = sorted(Path(td).glob("page-*.png"))
        if not pages:
            raise RuntimeError("pdftoppm produced no pages")
        applied = 0
        red_pages: list[Path] = []
        for idx, page_path in enumerate(pages, start=1):
            red_page = Path(td) / f"redacted-{idx}.png"
            page_regions = [
                r
                for r in regions
                if r["locator_type"] in {"page_normalized_bbox", "page_normalized_quad"}
                and int(r["locator"].get("page", idx)) == idx
            ]
            if page_regions:
                applied += redact_image(page_path, red_page, page_regions, page=idx)
            else:
                shutil.copy2(page_path, red_page)
            red_pages.append(red_page)
        if applied == 0:
            raise ValueError("no PDF page regions applied")
        out.parent.mkdir(parents=True, exist_ok=True)
        first = Image.open(red_pages[0])
        c = canvas.Canvas(str(out), pagesize=(first.width * 72 / dpi, first.height * 72 / dpi))
        for p in red_pages:
            img = Image.open(p)
            c.setPageSize((img.width * 72 / dpi, img.height * 72 / dpi))
            c.drawImage(ImageReader(img), 0, 0, width=img.width * 72 / dpi, height=img.height * 72 / dpi)
            c.showPage()
        c.save()
        return applied


def copy_redacted_payload(patient_dir: Path, entry: dict, out: Path) -> int:
    payload = entry.get("redacted_payload_path")
    if not payload:
        return 0
    src = patient_dir / safe_rel(payload)
    if not src.is_file():
        raise FileNotFoundError(f"redacted payload not found: {payload}")
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, out)
    return 1


def redact_docx(patient_dir: Path, source: Path, out: Path, entry: dict) -> int:
    copied = copy_redacted_payload(patient_dir, entry, out)
    if copied:
        return copied
    regions = [r for r in entry.get("regions", []) if r["locator_type"] == "xml_path"]
    if not regions:
        raise ValueError("DOCX requires xml_path regions or redacted_payload_path")
    out.parent.mkdir(parents=True, exist_ok=True)
    applied = 0
    with tempfile.TemporaryDirectory(prefix="cb_docx_") as td:
        root = Path(td)
        with zipfile.ZipFile(source) as z:
            z.extractall(root)
        by_part: dict[str, list[dict]] = {}
        for r in regions:
            part = r["locator"].get("part")
            if not isinstance(part, str) or not part.startswith("word/") or ".." in Path(part).parts:
                raise ValueError(f"unsafe docx part locator: {part!r}")
            by_part.setdefault(part, []).append(r)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for part, regs in by_part.items():
            xml_path = root / part
            if not xml_path.is_file():
                raise FileNotFoundError(f"DOCX XML part not found: {part}")
            tree = ET.parse(xml_path)
            texts = tree.findall(".//w:t", ns)
            for r in regs:
                loc = r["locator"]
                if "text_node_index" not in loc:
                    raise ValueError("xml_path locator requires text_node_index")
                idx = int(loc["text_node_index"])
                if idx < 0 or idx >= len(texts):
                    raise IndexError(f"text_node_index out of range: {idx}")
                texts[idx].text = MASK
                applied += 1
            tree.write(xml_path, encoding="utf-8", xml_declaration=True)
        with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
            for p in root.rglob("*"):
                if p.is_file():
                    z.write(p, p.relative_to(root).as_posix())
    return applied


def redact_xlsx(patient_dir: Path, source: Path, out: Path, entry: dict) -> int:
    copied = copy_redacted_payload(patient_dir, entry, out)
    if copied:
        return copied
    from openpyxl import load_workbook

    wb = load_workbook(source)
    applied = 0
    for r in entry.get("regions", []):
        if r["locator_type"] != "cell":
            continue
        loc = r["locator"]
        sheet = loc.get("sheet")
        cell = loc.get("cell")
        if sheet not in wb.sheetnames or not cell:
            raise ValueError(f"bad cell locator: {loc!r}")
        wb[sheet][cell].value = MASK
        if wb[sheet][cell].comment:
            wb[sheet][cell].comment = None
        applied += 1
    if applied == 0:
        raise ValueError("XLSX requires cell regions or redacted_payload_path")
    wb.properties.title = None
    wb.properties.subject = None
    wb.properties.creator = None
    wb.properties.lastModifiedBy = None
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return applied


def redact_text(patient_dir: Path, source: Path, out: Path, entry: dict) -> int:
    copied = copy_redacted_payload(patient_dir, entry, out)
    if copied:
        return copied
    text = source.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines(keepends=True)
    applied = 0
    for r in sorted((r for r in entry.get("regions", []) if r["locator_type"] == "line_span"), key=lambda x: (int(x["locator"].get("line", 0)), int(x["locator"].get("start", 0))), reverse=True):
        loc = r["locator"]
        line_no = int(loc["line"])
        if line_no < 1 or line_no > len(lines):
            raise IndexError(f"line locator out of range: {line_no}")
        line = lines[line_no - 1]
        newline = ""
        if line.endswith("\r\n"):
            body, newline = line[:-2], "\r\n"
        elif line.endswith("\n"):
            body, newline = line[:-1], "\n"
        else:
            body = line
        start = int(loc.get("start", 0))
        end = int(loc.get("end", len(body)))
        start = max(0, min(len(body), start))
        end = max(start, min(len(body), end))
        if end == start:
            raise ValueError(f"empty line span: {loc!r}")
        lines[line_no - 1] = body[:start] + MASK + body[end:] + newline
        applied += 1
    if applied == 0:
        raise ValueError("text-like source requires line_span regions or redacted_payload_path")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("".join(lines), encoding="utf-8")
    return applied


def redact_archive(patient_dir: Path, source: Path, out: Path, entry: dict) -> int:
    copied = copy_redacted_payload(patient_dir, entry, out)
    if copied:
        return copied
    raise ValueError("archive redaction requires redacted_payload_path rebuilt from redacted children")


def prepare_one(patient_dir: Path, entry: dict, previous: dict | None) -> dict:
    if previous and previous.get("status") == "done":
        return previous
    status = empty_entry(entry)
    source_abs = patient_dir / safe_rel(entry["bucket_path"])
    if not source_abs.is_file():
        status.update(status="failed", reason=f"bucket source missing: {entry['bucket_path']}")
        return status
    cand_rel = candidate_rel(entry, source_abs)
    cand_abs = patient_dir / cand_rel
    try:
        kind = entry["source_kind"]
        if kind == "image":
            applied = redact_image(source_abs, cand_abs, entry.get("regions", []))
        elif kind == "pdf":
            applied = redact_pdf(source_abs, cand_abs, entry.get("regions", []), entry.get("adapter_frame") or {})
        elif kind == "docx":
            applied = redact_docx(patient_dir, source_abs, cand_abs, entry)
        elif kind == "xlsx":
            applied = redact_xlsx(patient_dir, source_abs, cand_abs, entry)
        elif kind in {"csv", "txt", "html", "md"}:
            applied = redact_text(patient_dir, source_abs, cand_abs, entry)
        elif kind == "archive":
            applied = redact_archive(patient_dir, source_abs, cand_abs, entry)
        else:
            raise ValueError("unsupported source_kind")
        if applied <= 0:
            raise ValueError("no redaction regions applied")
        if sha256_file(source_abs) == sha256_file(cand_abs):
            raise ValueError("candidate hash equals source hash; refusing no-op redaction")
        status.update(
            status="redacted_pending_qa",
            redacted_candidate_path=cand_rel,
            coverage_passed=True,
            original_deleted=False,
            reason=None,
        )
    except Exception as e:
        status.update(
            status="blocked" if entry["source_kind"] == "other" else "failed",
            redacted_candidate_path=cand_rel if cand_abs.is_file() else None,
            coverage_passed=False,
            original_deleted=False,
            reason=f"{type(e).__name__}: {e}",
        )
    return status


def load_qa_report(path: Path) -> tuple[str, dict[str, bool]]:
    doc = read_json(path)
    report_id = doc.get("report_id") or doc.get("id") or path.stem
    decisions = {}
    for item in doc.get("files", []):
        fid = item.get("id")
        passed = item.get("pass")
        if fid and isinstance(passed, bool):
            decisions[fid] = passed
    return str(report_id), decisions


def commit_one(patient_dir: Path, entry: dict, status: dict, report_id: str, qa: dict[str, bool]) -> dict:
    if status.get("status") == "done":
        return status
    fid = entry["id"]
    passed = qa.get(fid)
    status["qa_report_id"] = report_id
    status["llm_qa_passed"] = passed
    status["qa_passed"] = bool(status.get("coverage_passed") is True and passed is True)
    if status["qa_passed"] is not True:
        status.update(status="failed", original_deleted=False, reason="LLM QA missing or failed")
        return status
    cand_rel = status.get("redacted_candidate_path")
    if not cand_rel:
        status.update(status="failed", original_deleted=False, reason="missing redacted_candidate_path")
        return status
    cand_abs = patient_dir / safe_rel(cand_rel)
    bucket_abs = patient_dir / safe_rel(entry["bucket_path"])
    mirror_abs = patient_dir / safe_rel(entry["mirror_path"])
    if not cand_abs.is_file():
        status.update(status="failed", original_deleted=False, reason=f"candidate missing: {cand_rel}")
        return status
    try:
        bucket_abs.parent.mkdir(parents=True, exist_ok=True)
        mirror_abs.parent.mkdir(parents=True, exist_ok=True)
        os.replace(cand_abs, bucket_abs)
        shutil.copy2(bucket_abs, mirror_abs)
        status.update(
            status="done",
            redacted_path=entry["bucket_path"],
            original_deleted=True,
            reason=None,
        )
    except Exception as e:
        status.update(status="failed", original_deleted=False, reason=f"commit failed: {type(e).__name__}: {e}")
    return status


def source_inventory_map(patient_dir: Path) -> dict[str, dict]:
    path = patient_dir / SOURCE_INVENTORY_NAME
    if not path.is_file():
        return {}
    try:
        doc = read_json(path)
        return {e["source_id"]: e for e in doc.get("files", []) if isinstance(e, dict) and e.get("source_id")}
    except Exception:
        return {}


def sync_source_status(patient_dir: Path, manifest: dict, status_entries: list[dict]) -> None:
    inv = source_inventory_map(patient_dir)
    by_file = {e["id"]: e for e in manifest.get("files", [])}
    out = []
    for st in status_entries:
        mf = by_file.get(st["id"], {})
        source_id = st["source_id"]
        inv_entry = inv.get(source_id, {})
        strategy = inv_entry.get("redaction_strategy") or STRATEGY_BY_KIND.get(mf.get("source_kind"), "blocked_unsupported")
        source_status = st["status"]
        if source_status == "redacted_pending_qa":
            source_status = "pending"
        out.append(
            {
                "source_id": source_id,
                "status": source_status,
                "strategy": strategy,
                "redacted_path": st.get("redacted_path"),
                "qa_passed": st.get("qa_passed"),
                "coverage_passed": st.get("coverage_passed"),
                "llm_qa_passed": st.get("llm_qa_passed"),
                "qa_report_id": st.get("qa_report_id"),
                "original_deleted": st.get("original_deleted"),
                "reason": st.get("reason"),
                "linked_redaction_manifest_id": st["id"],
            }
        )
    summary = {"total": len(out), "pending": 0, "done": 0, "failed": 0, "blocked": 0, "not_required": 0}
    for e in out:
        if e["status"] in summary:
            summary[e["status"]] += 1
    write_json(
        patient_dir / SOURCE_STATUS_NAME,
        {
            "schema": "source_redaction_status_v1",
            "patient_dir": str(patient_dir),
            "updated_at": now_iso(),
            "summary": summary,
            "files": out,
        },
    )


def run_prepare(args) -> int:
    _, manifest, patient_dir = resolve_manifest(args)
    previous = load_status(patient_dir)
    entries = []
    for entry in manifest.get("files", []):
        entries.append(prepare_one(patient_dir, entry, previous.get(entry["id"])))
    write_status(patient_dir, entries)
    sync_source_status(patient_dir, manifest, entries)
    return 1 if any(e["status"] in {"failed", "blocked"} for e in entries) else 0


def run_commit(args) -> int:
    _, manifest, patient_dir = resolve_manifest(args)
    qa_path = Path(args.qa_report or (patient_dir / "llm_redaction_qa.json")).resolve()
    if not qa_path.is_file():
        raise ValueError(f"QA report not found: {qa_path}")
    report_id, qa = load_qa_report(qa_path)
    previous = load_status(patient_dir)
    entries = []
    for entry in manifest.get("files", []):
        st = previous.get(entry["id"]) or empty_entry(entry)
        entries.append(commit_one(patient_dir, entry, st, report_id, qa))
    write_status(patient_dir, entries)
    sync_source_status(patient_dir, manifest, entries)
    return 1 if any(e["status"] != "done" for e in entries) else 0


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    sub = ap.add_subparsers(dest="command", required=True)
    for name in ("prepare", "commit"):
        sp = sub.add_parser(name)
        sp.add_argument("patient_dir", nargs="?")
        sp.add_argument("--manifest")
        if name == "commit":
            sp.add_argument("--qa-report")
    args = ap.parse_args(argv[1:])
    try:
        if args.command == "prepare":
            return run_prepare(args)
        if args.command == "commit":
            return run_commit(args)
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 2
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
